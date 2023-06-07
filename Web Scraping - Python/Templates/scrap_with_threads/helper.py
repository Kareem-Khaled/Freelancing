import re
import time
import openpyxl
import threading
from openpyxl.styles import Font
from openpyxl.utils import column_index_from_string

class Methods:
    def __init__(self, sheet_name, max_wr=5, slp=1, max_buffer=10):
        self.indices = {}
        self.max_wr = max_wr
        self.slp = slp
        self.row = 2
        self.idx = 0
        self.buffer = 0
        self.workbook = None
        self.worksheet = None
        self.writer_lock = threading.Lock()
        self.url_lock = threading.Lock()
        self.idx_lock = threading.Lock()
        self.bold_font = Font(bold=True)
        self.sheet_name = sheet_name
        self.urls = set()
        self.last_data = set()
        self.colindex = None
        self.max_buffer = max_buffer
        self.init_workbook()

    def init_workbook(self):
        try:
            self.workbook = openpyxl.load_workbook(self.sheet_name)
            self.worksheet = self.workbook.active
        except:
            self.workbook = openpyxl.Workbook()
            self.worksheet = self.workbook.active
            return

        for col in self.worksheet.iter_cols(min_row=1, max_row=1):
            for cell in col:
                if cell.value == 'Link':
                    self.colindex = column_index_from_string(cell.column_letter)

        if not self.colindex:
            return

        for ro in self.worksheet.iter_rows(min_row=2, min_col=self.colindex, max_col=self.colindex):
            for cell in ro:
                self.last_data.add(cell.value)

        self.row += len(self.last_data)

    def get_idx(self, x):
        with self.idx_lock:
            if x in self.indices:
                return self.indices[x]

            self.idx += 1
            self.indices[x] = self.idx
            return self.idx

    def last_save(self):
        if not self.buffer:
            return

        while True:
            try:
                self.workbook.save(self.sheet_name)
                print('Saving...', flush=True, end='')
                break
            except:
                time.sleep(self.slp)

    def write(self, id, data):
        try:
            print(f'[{id}]', flush=True, end='.')
            with self.writer_lock:
                for key, val in data:
                    col = self.get_idx(key)
                    alpha_key = re.match('[a-zA-Z\s]+', key).group()

                    wr = 0
                    while wr < self.max_wr:
                        try:
                            cell = self.worksheet.cell(1, col, alpha_key)
                            break
                        except:
                            wr += 1
                            time.sleep(self.slp * wr)

                    cell.font = self.bold_font
                    self.worksheet.cell(self.row, col, str(val))

                self.row += 1
                self.buffer += 1
                if self.buffer >= self.max_buffer:
                    self.buffer = 0
                    wr = 0
                    while wr < self.max_wr:
                        try:
                            self.workbook.save(self.sheet_name)
                            print('Saving...', flush=True, end='')
                            break
                        except:
                            wr += 1
                            time.sleep(self.slp * wr)

        except Exception as e:
            print("- Error in writing")
            print(str(e))

